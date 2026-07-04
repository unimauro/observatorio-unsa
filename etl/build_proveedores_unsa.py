import openpyxl, json, os
HERE=os.path.dirname(os.path.abspath(__file__))
CONOSCE=os.environ.get('CONOSCE_DIR','/Users/unimauro/Documents/Repos/observatorio-sanmarcos/etl/conosce')
CODIGO='001908'; RUC_ENT='20163646499'  # UNSA - pliego 513
def tp(ruc):
    s=str(ruc)
    # consorcio/codigo SEACE: no 11 digitos o no empieza en 10/15/20 -> empresa
    if len(s)!=11 or s[:2] not in ('10','15','20'):
        return 'empresa'
    return 'natural' if s[:2] in ('10','15') else 'empresa'

# dueno/representante (via datosperu.org / reporte consorcios CONOSCE). null = sin fuente confiable.
DP='https://www.datosperu.org/'
CONS='https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/consorcios/'
duenos={
 # empresas (RUC 20) - gerente general / apoderado segun datosperu.org
 '20100147514':('Sucursal de Southern Copper Corporation, controlada por Grupo Mexico (German Larrea Mota-Velasco)', DP+'empresa-southern-peru-copper-corporation-sucursal-del-peru-20100147514.php'),
 '20603237120':('Gerente general: Samanez Alarcon Rodolfo Enrique', DP+'empresa-equipos-analiticos-y-tecnologia-de-informacion-sac-equanti-sac-20603237120.php'),
 '20600229762':('Apoderados: Barrios Fernandez Concha Raul Hernando Martin; Bolaños Salazar Milton Cesar. Sucursal de Rohde & Schwarz (Alemania)', DP+'empresa-rohde-schwarz-colombia-sas-sucursal-peru-20600229762.php'),
 '20563651921':('Gerente general: Silvia Principe Salazar', DP+'empresa-cima-book-sac-20563651921.php'),
 '20606335581':('Gerente general: Martinez Machuca Alejandra Soledad', DP+'empresa-shopmarket-sac-20606335581.php'),
 '20512011919':('Gerente general: Del Aguila Herrera Ricardo', DP+'empresa-proyecta-ingenieros-civiles-sac-20512011919.php'),
 '20521301237':('Gerente general: Sotomayor Torres Andy Ronal', DP+'empresa-warem-sac-20521301237.php'),
 '20178285336':('Gerente general: Lay Whittembury Christopher', DP+'empresa-reactivos-para-analisis-sac-20178285336.php'),
 '20602629806':('Gerente general: Aza Eyzaguirre Fredy Oscar', DP+'empresa-ingenieros-civiles-aza-sarmiento-yanqui-asociados-sociedad-anonima-cerrada-icaza-sac-20602629806.php'),
 '20392986601':('Gerente general: Franco Cruz Aaron', DP+'empresa-beicon-contratistas-generales-sac-beicon-sac-20392986601.php'),
 '20501610217':('Gerente: Morales Figueroa Juan Zenon; apoderado: Morales Rodriguez Giancarlo', DP+'empresa-g-f-cueros-y-derivados-sociedad-anonima-cerrada-g-f-cueros-y-derivados-sac-20501610217.php'),
 # consorcios (codigo SEACE) - integrantes y sus gerentes generales
 '579383':('Consorcio integrado por CAMPUS CONTRATISTAS GENERALES S.R.L. (RUC 20454146477, gte. Benavente Barreda Andres Bertin) y MAGNUS CONTRATISTAS GENERALES S.A.C. (RUC 20498467971, gte. Sucasaca Caceres Angela Maria)', CONS+'2023/CONOSCE_CONSORCIO2023_0.xlsx'),
 '1673972':('Consorcio integrado por CONSTRUCTORA D & R S.A.C. (RUC 20448302211, gte. Coyla Apaza Reynaldo) y GTECH S.A.C. (RUC 20448000711, gte. Guerra Bueno Alan Nahel)', CONS+'2025/CONOSCE_CONSORCIO2025_0.xlsx'),
 '1529121':('Consorcio integrado por GRUPO FER. CONS S.A.C. (RUC 20447689503, gte. Chambilla Chaparro Sabino Freddy) y ARCEN CONTRATISTAS GENERALES S.A.C. (RUC 20455430489, gte. Condori Mamani Ronald)', CONS+'2024/CONOSCE_CONSORCIO2024_0.xlsx'),
}

rows=[]
for Y in (2023,2024,2025):
    wb=openpyxl.load_workbook(f"{CONOSCE}/CONOSCE_ADJUDICACIONES{Y}_0.xlsx", read_only=True)
    ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); next(it)
    for r in it:
        if str(r[0])==CODIGO and str(r[1])==RUC_ENT:
            rows.append(r)
    wb.close()
print("item rows:", len(rows))

agg={}
for r in rows:
    ruc=str(r[19]); monto=r[15] or 0
    a=agg.setdefault(ruc,{'nombre':r[20],'ruc':ruc,'monto':0.0,'convs':set(),
        'tipos':{},'objeto':{},'tipo_persona':tp(ruc),'tipo_proveedor_seace':r[21]})
    a['monto']+=float(monto)
    a['convs'].add(r[5])
    a['tipos'][r[7]]=a['tipos'].get(r[7],0)+1
    a['objeto'][r[6]]=a['objeto'].get(r[6],0)+1
provs=[]
for ruc,a in agg.items():
    d,f=duenos.get(ruc,(None,None))
    provs.append({'nombre':a['nombre'],'ruc':ruc,'monto':round(a['monto'],2),
        'n':len(a['convs']),'tipos':a['tipos'],'objeto':a['objeto'],
        'tipo_persona':a['tipo_persona'],'tipo_proveedor_seace':a['tipo_proveedor_seace'],
        'dueno':d,'fuente_dueno':f})
provs.sort(key=lambda x:-x['monto'])
monto_total=round(sum(p['monto'] for p in provs),2)
emp=[p for p in provs if p['tipo_persona']=='empresa']
nat=[p for p in provs if p['tipo_persona']=='natural']
all_convs=set(r[5] for r in rows)
top_personas=sorted([{'nombre':p['nombre'],'ruc':p['ruc'],'monto':p['monto'],'n':p['n']} for p in nat],key=lambda x:-x['monto'])[:20]
out={
 '_meta':{
  'fuente':'OECE/OSCE - CONOSCE Datos Abiertos, reporte de Adjudicaciones (buena pro por item)',
  'fuente_url':'https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/adjudicaciones/',
  'entidad':'Universidad Nacional de San Agustin de Arequipa (UNSA) - pliego 513',
  'ruc':RUC_ENT,
  'codigoentidad_conosce':int(CODIGO),
  'periodo':'2023-2025',
  'extraido':'2026-07',
  'unidad_monto':'Soles (PEN), monto adjudicado por item',
  'nota':"Agregado desde reportes anuales CONOSCE de Adjudicaciones (nivel item de buena pro), archivos CONOSCE_ADJUDICACIONES{2023,2024,2025}_0.xlsx, filtrando codigoentidad 001908 / RUC 20163646499 (UNIVERSIDAD NACIONAL SAN AGUSTIN - UNSA, Arequipa, pliego 513). 'monto' = suma de monto_adjudicado_item_soles. 'n' = numero de procesos (codigoconvocatoria) distintos en que ese proveedor obtuvo buena pro. 'tipos' = procesos por tipo de proceso de seleccion. 'objeto' = items por objeto contractual (Bien/Servicio/Obra). tipo_persona por prefijo de RUC (20=empresa, 10/15=natural); consorcios (codigo SEACE, no RUC de 11 digitos 10/15/20) clasificados como empresa. NO incluye ordenes de compra <8 UIT (dataset aparte)."
 },
 'totales':{
  'monto_total':monto_total,
  'n_proveedores':len(provs),
  'n_procesos':len(all_convs),
  'n_empresas':len(emp),
  'n_personas_naturales':len(nat),
  'monto_empresas':round(sum(p['monto'] for p in emp),2),
  'monto_personas_naturales':round(sum(p['monto'] for p in nat),2),
 },
 'top_personas':top_personas,
 'proveedores':provs,
}
json.dump(out, open(os.path.join(HERE,'..','data','proveedores-unsa.json'),'w'),
          ensure_ascii=False, separators=(",",":"))
print("proveedores:", len(provs), "monto_total:", monto_total, "procesos:", len(all_convs))
print("empresas:", len(emp), "naturales:", len(nat))
print("--- TOP 20 ---")
for p in provs[:20]:
    print(f"  {p['monto']:>14,.2f}  {p['ruc']:12}  {p['tipo_persona']:8}  {p['nombre']}")
